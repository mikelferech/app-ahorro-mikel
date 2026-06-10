import base64
import calendar
import re
from datetime import date, datetime, timedelta
from pathlib import Path

import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from openpyxl import load_workbook

APP_TITLE = "Ahorro Mikel"
APP_VERSION = "0.4.0"
APP_UPDATED = "10/06/2026"
EXCEL_PATH = Path("Ahorro.xlsx")
BUDGET_PATH = Path("Presupuesto_Casa.xlsx")
ASSETS = Path("assets")
MFE_LOGO = ASSETS / "mfe_cabecera.png"
VADILLO_LOGO = ASSETS / "vadillo.svg"
MONTHS_ES = ["ENE", "FEB", "MAR", "ABR", "MAY", "JUN", "JUL", "AGO", "SEP", "OCT", "NOV", "DIC"]
DEFAULT_BANKS = ["BBVA", "Openbank", "Cajamar", "Otros"]
RETENCION_INTERESES = 0.19
VACACIONES_ANUALES = 23

st.set_page_config(page_title=APP_TITLE, page_icon="💰", layout="wide", initial_sidebar_state="collapsed")

st.markdown("""
<style>
.main .block-container {padding-top: 1.25rem; max-width: 1500px;}
div[data-testid="stMetricValue"] {font-size: 2.0rem;}
.login-wrap {max-width: 460px; margin: 4.5rem auto 0 auto; padding: 1.6rem 1.6rem 1.2rem 1.6rem; border-radius: 20px; border: 1px solid rgba(128,128,128,.22); background: rgba(128,128,128,.06); text-align:center;}
.login-wrap img {max-width: 260px; margin-bottom: 1rem;}
.header-brand {display:flex; align-items:center; gap:18px; margin-bottom: .25rem;}
.header-brand img {max-height:64px; max-width:230px; object-fit:contain;}
.vadillo-box {background:#ffffff; border-radius:18px; padding:14px; display:flex; align-items:center; justify-content:center; margin-bottom:16px; border:1px solid rgba(0,0,0,.08);}
.vadillo-box img {max-height:90px; max-width:95%; object-fit: contain;}
@media (prefers-color-scheme: dark) {
  .vadillo-box {background:#111827; border-color:rgba(255,255,255,.12);}
  .vadillo-box img {filter: grayscale(1) brightness(0) invert(1);}
}
.bank-chip {border-radius:10px; padding:8px 12px; font-weight:800; text-align:center; color:#fff; margin-bottom:8px;}
.cal-grid {display:grid; grid-template-columns: repeat(7, 1fr); gap:4px; margin-bottom:18px;}
.cal-head {text-align:center; font-weight:800; font-size:.78rem; opacity:.75;}
.cal-day {min-height:34px; border-radius:7px; padding:5px; text-align:center; font-size:.82rem; border:1px solid rgba(128,128,128,.18);}
.cal-empty {opacity:.15;}
.cal-red {background:#b91c1c; color:#fff;}
.cal-maroon {background:#5b1720; color:#fff;}
.cal-grey {background:#6b7280; color:#fff;}
.cal-green {background:#15803d; color:#fff; font-weight:800;}
.cal-normal {background:rgba(128,128,128,.09);}
.footer {margin-top:2.5rem; padding:1.2rem 0 .4rem 0; border-top:1px solid rgba(128,128,128,.22); opacity:.82; font-size:.86rem; display:flex; gap:12px; align-items:center; justify-content:center; flex-wrap:wrap;}
.footer img {height:28px; width:auto;}
.small-note {font-size:.85rem; opacity:.78;}
</style>
""", unsafe_allow_html=True)

# ---------- utilidades ----------
def img_b64(path: Path) -> str:
    try:
        return base64.b64encode(path.read_bytes()).decode()
    except Exception:
        return ""

def image_html(path: Path) -> str:
    b64 = img_b64(path)
    if not b64: return ""
    ext = path.suffix.lower().replace(".", "") or "png"
    if ext == "jpg": ext = "jpeg"
    if ext == "svg": ext = "svg+xml"
    return f"data:image/{ext};base64,{b64}"

def money(v) -> float:
    if v is None or v == "": return 0.0
    if isinstance(v, (int, float)): return float(v)
    if isinstance(v, str):
        s = v.replace("€", "").replace(" ", "").strip()
        if "," in s and "." in s:
            s = s.replace(".", "").replace(",", ".")
        elif "," in s:
            s = s.replace(",", ".")
        try: return float(s)
        except Exception: return 0.0
    return 0.0

def euro(v) -> str:
    try:
        s = f"{float(v):,.2f} €"
        return s.replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "0,00 €"

def mes_label(d) -> str:
    if pd.isna(d): return ""
    if not isinstance(d, (pd.Timestamp, datetime, date)):
        d = pd.to_datetime(d, errors="coerce")
    if pd.isna(d): return ""
    return f"{MONTHS_ES[d.month-1]}{str(d.year)[-2:]}"

def normalize_month(d) -> date:
    if isinstance(d, datetime): d = d.date()
    if isinstance(d, pd.Timestamp): d = d.date()
    return date(d.year, d.month, 1)

def date_range(a: date, b: date):
    if isinstance(a, datetime): a = a.date()
    if isinstance(b, datetime): b = b.date()
    cur = a
    while cur <= b:
        yield cur
        cur += timedelta(days=1)

def safe_key(name: str) -> str:
    s = re.sub(r"[^A-Za-z0-9]+", "_", str(name).strip()).strip("_")
    return s or "Banco"

def month_options(min_year=None, max_year=None):
    today = date.today()
    min_year = min_year or 2012
    max_year = max_year or today.year + 8
    out=[]
    for y in range(int(min_year), int(max_year)+1):
        for m in range(1,13):
            d=date(y,m,1); out.append((d, mes_label(d)))
    return out

def first_day_for_label(label: str) -> date:
    label = str(label).upper().strip()
    m = MONTHS_ES.index(label[:3]) + 1
    y = 2000 + int(label[3:])
    return date(y,m,1)

# ---------- excel ----------
def wb_data_only(): return load_workbook(EXCEL_PATH, data_only=True)
def wb_write(): return load_workbook(EXCEL_PATH)

def sheet_to_df(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return pd.DataFrame()
    headers = [str(h) if h is not None else f"Col{i}" for i,h in enumerate(rows[0],1)]
    return pd.DataFrame(rows[1:], columns=headers).dropna(how="all")

def replace_sheet(wb, title, df):
    if title in wb.sheetnames: del wb[title]
    ws = wb.create_sheet(title)
    ws.append(list(df.columns))
    for _, row in df.iterrows():
        vals=[]
        for c in df.columns:
            v = row.get(c)
            if isinstance(v, pd.Timestamp): v = v.date()
            vals.append(v)
        ws.append(vals)
    return ws

def load_table(sheet, headers) -> pd.DataFrame:
    if not EXCEL_PATH.exists(): return pd.DataFrame(columns=headers)
    wb = wb_data_only()
    if sheet not in wb.sheetnames: return pd.DataFrame(columns=headers)
    df = sheet_to_df(wb[sheet])
    if df.empty: return pd.DataFrame(columns=headers)
    for h in headers:
        if h not in df: df[h] = None
    return df[headers]

def save_table(sheet, df):
    wb = wb_write()
    replace_sheet(wb, sheet, df)
    wb.save(EXCEL_PATH)
    st.cache_data.clear()

# ---------- auth ----------
def get_secret_auth():
    try: return st.secrets["auth"]["username"], st.secrets["auth"]["password"]
    except Exception: return None, None

def render_logo(path, max_width=280):
    src = image_html(path)
    if src:
        st.markdown(f'<img src="{src}" style="max-width:{max_width}px;width:75%;height:auto;">', unsafe_allow_html=True)

def login_gate():
    user, password = get_secret_auth()
    if "auth_ok" not in st.session_state: st.session_state.auth_ok = False
    if st.session_state.auth_ok: return
    _, center, _ = st.columns([1, 1.15, 1])
    with center:
        src = image_html(MFE_LOGO)
        logo = f'<img src="{src}">' if src else ''
        st.markdown(f'<div class="login-wrap">{logo}<h2>🔒 Ahorro Mikel</h2><p>Acceso privado</p></div>', unsafe_allow_html=True)
        if not user or not password:
            st.error("Faltan credenciales en Streamlit Secrets.")
            st.code('[auth]\nusername = "mikelferech"\npassword = "TU_CONTRASEÑA"')
            st.stop()
        u = st.text_input("Usuario", key="login_user")
        p = st.text_input("Contraseña", type="password", key="login_pass")
        if st.button("Entrar", use_container_width=True, key="login_btn"):
            if u == user and p == password:
                st.session_state.auth_ok = True
                st.rerun()
            else:
                st.error("Usuario o contraseña incorrectos")
    st.stop()

# ---------- bancos ----------
def default_bank_config():
    return pd.DataFrame([
        {"Clave":"BBVA", "Nombre":"BBVA", "Color":"#072146", "Activo":True, "Orden":1},
        {"Clave":"Openbank", "Nombre":"Openbank", "Color":"#e40046", "Activo":True, "Orden":2},
        {"Clave":"Cajamar", "Nombre":"Cajamar", "Color":"#008c95", "Activo":True, "Orden":3},
        {"Clave":"Otros", "Nombre":"Otros", "Color":"#6b7280", "Activo":True, "Orden":4},
    ])

def boolify(x):
    if isinstance(x, bool): return x
    return str(x).strip().lower() not in ("false","0","no","n","")

def load_bank_config() -> pd.DataFrame:
    headers=["Clave","Nombre","Color","Activo","Orden"]
    df = load_table("App_Bancos", headers)
    if df.empty:
        df = default_bank_config()
        save_table("App_Bancos", df)
    for h in headers:
        if h not in df: df[h]=None
    df["Clave"] = df["Clave"].fillna("").astype(str).apply(safe_key)
    df["Nombre"] = df["Nombre"].fillna(df["Clave"]).astype(str)
    df["Color"] = df["Color"].fillna("#6b7280").astype(str)
    df["Activo"] = df["Activo"].apply(boolify)
    df["Orden"] = pd.to_numeric(df["Orden"], errors="coerce").fillna(99).astype(int)
    df = df.drop_duplicates("Clave", keep="first")
    return df.sort_values("Orden")[headers]

def all_bank_keys(include_inactive=True):
    cfg=load_bank_config()
    if not include_inactive: cfg=cfg[cfg["Activo"]]
    return cfg["Clave"].tolist()

def bank_name(clave: str):
    cfg=load_bank_config(); row=cfg[cfg["Clave"]==clave]
    return str(row.iloc[0]["Nombre"]) if not row.empty else clave

def bank_color(clave: str):
    cfg=load_bank_config(); row=cfg[cfg["Clave"]==clave]
    return str(row.iloc[0]["Color"]) if not row.empty else "#6b7280"

def render_bank_config(location="global"):
    with st.expander("⚙️ Configurar bancos / colores", expanded=False):
        st.caption("Añade bancos, cambia nombres, colores, orden u ocúltalos sin perder datos históricos.")
        cfg=load_bank_config().copy()
        edited=st.data_editor(
            cfg,
            use_container_width=True,
            hide_index=True,
            num_rows="dynamic",
            key=f"bank_config_editor_{location}",
            column_config={
                "Clave": st.column_config.TextColumn("Clave interna"),
                "Nombre": st.column_config.TextColumn("Nombre visible"),
                "Color": st.column_config.TextColumn("Color HEX"),
                "Activo": st.column_config.CheckboxColumn("Mostrar"),
                "Orden": st.column_config.NumberColumn("Orden", step=1),
            },
        )
        c1,c2=st.columns(2)
        if c1.button("Guardar bancos", use_container_width=True, key=f"save_bank_config_{location}"):
            clean=edited.copy()
            clean["Clave"] = clean.apply(lambda r: safe_key(r.get("Clave") or r.get("Nombre")), axis=1)
            clean["Nombre"] = clean["Nombre"].fillna(clean["Clave"])
            clean["Color"] = clean["Color"].fillna("#6b7280")
            clean["Activo"] = clean["Activo"].apply(boolify)
            clean["Orden"] = pd.to_numeric(clean["Orden"], errors="coerce").fillna(range(1, len(clean)+1)).astype(int)
            clean = clean.drop_duplicates("Clave", keep="first")
            save_table("App_Bancos", clean[["Clave","Nombre","Color","Activo","Orden"]])
            st.success("Bancos guardados."); st.rerun()
        if c2.button("Restaurar bancos base", use_container_width=True, key=f"reset_bank_config_{location}"):
            save_table("App_Bancos", default_bank_config())
            st.success("Bancos base restaurados."); st.rerun()

# ---------- ahorro ----------
@st.cache_data(show_spinner=False)
def extract_original_ahorro() -> pd.DataFrame:
    if not EXCEL_PATH.exists(): return pd.DataFrame()
    wb = wb_data_only()
    if "Ahorro BBVA" not in wb.sheetnames: return pd.DataFrame()
    ws = wb["Ahorro BBVA"]
    rows=[]
    for r in range(1, ws.max_row+1):
        d = ws.cell(r,1).value
        if not isinstance(d, (datetime, date)): continue
        bbva = money(ws.cell(r,4).value if ws.cell(r,4).value not in (None,"") else ws.cell(r,2).value)
        open_caj = money(ws.cell(r,8).value)
        total = money(ws.cell(r,9).value)
        if total == 0: total = bbva + open_caj
        if total == 0 and bbva == 0 and open_caj == 0: continue
        rows.append({"Fecha": normalize_month(d), "BBVA":bbva, "Openbank":open_caj, "Cajamar":0.0, "Otros":0.0})
    df=pd.DataFrame(rows)
    if df.empty: return df
    return df.drop_duplicates("Fecha", keep="first").sort_values("Fecha")

def load_app_ahorro() -> pd.DataFrame:
    keys=all_bank_keys(True)
    headers=["Fecha"]+keys+["Total","+/-"]
    df=load_table("App_Ahorro", headers)
    if df.empty:
        df=extract_original_ahorro()
        if not df.empty: save_ahorro(df)
        return df
    df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce").dt.date
    df=df.dropna(subset=["Fecha"])
    df["Fecha"] = df["Fecha"].apply(normalize_month)
    for k in keys:
        if k not in df: df[k]=0.0
        df[k]=df[k].apply(money)
    df=df.drop_duplicates("Fecha", keep="last").sort_values("Fecha")
    df["Total"] = df[keys].sum(axis=1)
    df["+/-"] = df["Total"].diff().fillna(0)
    return df

def save_ahorro(df: pd.DataFrame):
    keys=all_bank_keys(True)
    out=df.copy()
    if out.empty:
        out=pd.DataFrame(columns=["Fecha"]+keys)
    for k in keys:
        if k not in out: out[k]=0.0
        out[k]=out[k].apply(money)
    out["Fecha"] = pd.to_datetime(out["Fecha"], errors="coerce").dt.date
    out = out.dropna(subset=["Fecha"])
    out["Fecha"] = out["Fecha"].apply(normalize_month)
    out = out.drop_duplicates("Fecha", keep="last").sort_values("Fecha")
    out["Total"] = out[keys].sum(axis=1)
    out["+/-"] = out["Total"].diff().fillna(0)
    save_table("App_Ahorro", out[["Fecha"]+keys+["Total","+/-"]])

def kpi_dashboard(df):
    if df.empty or "Total" not in df:
        st.warning("Aún no hay datos válidos de ahorro."); return
    valid=df[df["Total"].apply(money)>0].sort_values("Fecha")
    if valid.empty:
        st.warning("Aún no hay datos válidos de ahorro."); return
    last=valid.iloc[-1]; prev=valid.iloc[-2] if len(valid)>1 else None
    diff = last["Total"] - prev["Total"] if prev is not None else last.get("+/-",0)
    keys=all_bank_keys(False)
    c=st.columns(4)
    c[0].metric(f"Total actual · {mes_label(last['Fecha'])}", euro(last["Total"]))
    c[1].metric("Último +/-", euro(diff), delta=euro(diff))
    if keys:
        c[2].metric(bank_name(keys[0]), euro(last.get(keys[0],0)))
    if len(keys)>1:
        c[3].metric(" + ".join(bank_name(k) for k in keys[1:3]), euro(sum(money(last.get(k,0)) for k in keys[1:3])))

def dataframe_money(df, money_cols):
    out=df.copy()
    for col in money_cols:
        if col in out: out[col]=out[col].apply(euro)
    return out

def render_ahorro():
    st.header("💰 Evolución de ahorro")
    df=load_app_ahorro(); kpi_dashboard(df); render_bank_config("ahorro")
    keys=all_bank_keys(False)
    with st.expander("➕ Introducir / actualizar saldo mensual", expanded=False):
        today=date.today(); min_year=2012
        if not df.empty: min_year=min(d.year for d in df["Fecha"])
        opts=month_options(min_year, today.year+8)
        vals=[o[0] for o in opts]
        current=date(today.year,today.month,1)
        idx=vals.index(current) if current in vals else len(vals)-1
        selected=st.selectbox("Mes", vals, index=idx, format_func=mes_label, key="ahorro_mes")
        existing=df[df["Fecha"]==selected] if not df.empty else pd.DataFrame()
        defaults=existing.iloc[0].to_dict() if not existing.empty else {}
        cols=st.columns(max(1,len(keys)))
        row={"Fecha":selected}
        for i,k in enumerate(keys):
            row[k]=cols[i].number_input(f"Saldo {bank_name(k)}", value=float(money(defaults.get(k,0))), step=0.01, format="%.2f", key=f"saldo_{k}_{selected}")
        for k in all_bank_keys(True):
            if k not in row: row[k]=money(defaults.get(k,0))
        if st.button("Guardar saldo", use_container_width=True, key="save_saldo"):
            df2=df[df["Fecha"]!=selected].copy() if not df.empty else pd.DataFrame()
            df2=pd.concat([df2,pd.DataFrame([row])], ignore_index=True)
            save_ahorro(df2); st.success("Saldo guardado."); st.rerun()
    if df.empty: return
    df=df.sort_values("Fecha")
    min_d,max_d=df["Fecha"].min(),df["Fecha"].max()
    a,b=st.columns(2)
    start=a.selectbox("Mostrar desde", [o[0] for o in month_options(min_d.year,max_d.year+1)], index=0, format_func=mes_label, key="ah_start")
    end_opts=[o[0] for o in month_options(min_d.year,max_d.year+1)]
    end_index=max([i for i,d in enumerate(end_opts) if d<=max_d] or [len(end_opts)-1])
    end=b.selectbox("Mostrar hasta", end_opts, index=end_index, format_func=mes_label, key="ah_end")
    chart=df[(df["Fecha"]>=start)&(df["Fecha"]<=end)].copy().sort_values("Fecha")
    chart["Mes"]=chart["Fecha"].apply(mes_label)
    fig=go.Figure(go.Scatter(x=chart["Mes"], y=chart["Total"], mode="lines+markers", name="Total"))
    fig.update_layout(title="Evolución patrimonio", yaxis_title="€", height=410)
    st.plotly_chart(fig, use_container_width=True)
    colors=["#16a34a" if v>=0 else "#dc2626" for v in chart["+/-"]]
    fig2=go.Figure(go.Bar(x=chart["Mes"], y=chart["+/-"], marker_color=colors))
    fig2.update_layout(title="Diferencia con el mes anterior", yaxis_title="€", height=340)
    st.plotly_chart(fig2, use_container_width=True)
    st.subheader("Saldos por banco")
    chip_cols=st.columns(max(1,len(keys)))
    for i,k in enumerate(keys): chip_cols[i].markdown(f'<div class="bank-chip" style="background:{bank_color(k)}">{bank_name(k)}</div>', unsafe_allow_html=True)
    view=df.sort_values("Fecha", ascending=False).copy(); view["Mes"]=view["Fecha"].apply(mes_label)
    rename={k:bank_name(k) for k in keys}
    shown=view[["Mes"]+keys+["Total","+/-"]].rename(columns=rename)
    st.dataframe(dataframe_money(shown, list(rename.values())+["Total","+/-"]), use_container_width=True, hide_index=True)
    with st.expander("✏️ Editar o borrar líneas de ahorro", expanded=False):
        edit=df.sort_values("Fecha", ascending=False).copy(); edit["Borrar"]=False; edit["Mes"]=edit["Fecha"].apply(mes_label)
        cols=["Borrar","Mes"]+all_bank_keys(True)
        edited=st.data_editor(edit[cols], use_container_width=True, hide_index=True, key="editor_ahorro", column_config={"Borrar":st.column_config.CheckboxColumn("Borrar"), "Mes":st.column_config.TextColumn("Mes", disabled=True), **{k:st.column_config.NumberColumn(bank_name(k), step=0.01, format="%.2f") for k in all_bank_keys(True)}})
        if st.button("Guardar cambios", use_container_width=True, key="save_editor_ahorro"):
            base=edit[["Fecha","Mes"]].reset_index(drop=True); ed=edited.reset_index(drop=True)
            out=base.join(ed.drop(columns=["Mes"], errors="ignore"))
            out=out[~out["Borrar"].fillna(False)].drop(columns=["Borrar","Mes"], errors="ignore")
            save_ahorro(out); st.success("Cambios guardados."); st.rerun()

# ---------- nóminas/vacaciones/festivos ----------
def easter_date(year:int)->date:
    a=year%19; b=year//100; c=year%100; d=b//4; e=b%4; f=(b+8)//25; g=(b-f+1)//3
    h=(19*a+b-d-g+15)%30; i=c//4; k=c%4; l=(32+2*e+2*i-h-k)%7; m=(a+11*h+22*l)//451
    mo=(h+l-7*m+114)//31; da=((h+l-7*m+114)%31)+1
    return date(year,mo,da)

def festivos_vitoria_base(year:int)->pd.DataFrame:
    e=easter_date(year); rows=[]
    fixed=[(1,1,"Año Nuevo"),(1,6,"Reyes"),(4,28,"San Prudencio"),(5,1,"Día del Trabajo"),(7,25,"Santiago"),(8,5,"Virgen Blanca"),(8,15,"Asunción"),(10,12,"Fiesta Nacional"),(11,1,"Todos los Santos"),(12,6,"Constitución"),(12,8,"Inmaculada"),(12,25,"Navidad")]
    for m,d,n in fixed: rows.append({"Año":year,"Fecha":date(year,m,d).isoformat(),"Nombre":n,"Activo":True,"Origen":"Base"})
    rows += [{"Año":year,"Fecha":(e-timedelta(days=2)).isoformat(),"Nombre":"Viernes Santo","Activo":True,"Origen":"Base"},{"Año":year,"Fecha":(e+timedelta(days=1)).isoformat(),"Nombre":"Lunes de Pascua","Activo":True,"Origen":"Base"}]
    return pd.DataFrame(rows)

def load_festivos_df(year:int)->pd.DataFrame:
    headers=["Año","Fecha","Nombre","Activo","Origen"]
    saved=load_table("App_Festivos", headers)
    saved_y=saved[pd.to_numeric(saved.get("Año"), errors="coerce")==year].copy() if not saved.empty else pd.DataFrame(columns=headers)
    if saved_y.empty:
        base=festivos_vitoria_base(year)
        save_table("App_Festivos", pd.concat([saved, base], ignore_index=True) if not saved.empty else base)
        return base
    return saved_y[headers]

def save_festivos_year(year:int, df_year:pd.DataFrame):
    headers=["Año","Fecha","Nombre","Activo","Origen"]
    cur=load_table("App_Festivos", headers)
    other=cur[pd.to_numeric(cur.get("Año"), errors="coerce")!=year].copy() if not cur.empty else pd.DataFrame(columns=headers)
    clean=df_year.copy()
    for h in headers:
        if h not in clean: clean[h]=None
    clean["Año"]=year
    clean["Fecha"]=pd.to_datetime(clean["Fecha"], errors="coerce").dt.date.apply(lambda x: x.isoformat() if pd.notna(x) else "")
    clean=clean[clean["Fecha"]!=""]
    clean["Nombre"]=clean["Nombre"].fillna(""); clean["Activo"]=clean["Activo"].apply(boolify); clean["Origen"]=clean["Origen"].fillna("Manual")
    save_table("App_Festivos", pd.concat([other, clean[headers]], ignore_index=True))

def festivos_vitoria(year:int)->set:
    df=load_festivos_df(year); out=set()
    for _,r in df.iterrows():
        if not boolify(r.get("Activo", True)): continue
        try: out.add(pd.to_datetime(r["Fecha"]).date())
        except Exception: pass
    return out

def is_intensiva(day:date, festivos:set)->bool:
    return day.weekday()==4 or (day+timedelta(days=1) in festivos) or (date(day.year,6,1)<=day<=date(day.year,9,30))

def working_vacation_days(start:date,end:date,festivos:set)->int:
    return sum(1 for d in date_range(start,end) if d.weekday()<5 and d not in festivos)

def load_nominas():
    headers=["Año","Mes","Fecha","Bruto","SS","Desempleo","IRPF_pct","IRPF","Neto_calculado","Ingresado","Diferencia","Estado"]
    df=load_table("App_Nominas", headers)
    if not df.empty: return df
    rows=[]
    try:
        wb=wb_data_only()
        if "25 26 Vadillo" in wb.sheetnames:
            ws=wb["25 26 Vadillo"]
            for r in range(1, ws.max_row+1):
                d=ws.cell(r,3).value
                if not isinstance(d,(datetime,date)): continue
                bruto=money(ws.cell(r,4).value); ss=money(ws.cell(r,5).value); desemp=money(ws.cell(r,6).value); irpf=money(ws.cell(r,8).value); ing=money(ws.cell(r,10).value)
                if bruto==0 and ing==0: continue
                pct=irpf/bruto if bruto else .13; neto=bruto-ss-desemp-irpf
                rows.append({"Año":d.year,"Mes":d.month,"Fecha":normalize_month(d),"Bruto":bruto,"SS":ss,"Desempleo":desemp,"IRPF_pct":pct,"IRPF":irpf,"Neto_calculado":neto,"Ingresado":ing,"Diferencia":ing-neto if ing else 0,"Estado":""})
    except Exception: pass
    df=pd.DataFrame(rows, columns=headers)
    if not df.empty: save_table("App_Nominas", df)
    return df

def render_nominas():
    st.header("💼 Nóminas")
    src=image_html(VADILLO_LOGO)
    if src: st.markdown(f'<div class="vadillo-box"><img src="{src}" alt="Grupo Vadillo Asesores"></div>', unsafe_allow_html=True)
    df=load_nominas(); years=list(range(date.today().year-1, date.today().year+8))
    year=st.selectbox("Año", years, index=years.index(date.today().year), key="nom_year")
    month=st.selectbox("Mes", list(range(1,13)), index=date.today().month-1, format_func=lambda m:f"{MONTHS_ES[m-1]}{str(year)[-2:]}", key="nom_month")
    cur=df[(pd.to_numeric(df.get("Año"), errors="coerce")==year)&(pd.to_numeric(df.get("Mes"), errors="coerce")==month)] if not df.empty else pd.DataFrame()
    row=cur.iloc[0].to_dict() if not cur.empty else {}
    c1,c2,c3,c4,c5=st.columns(5)
    bruto=c1.number_input("Bruto", value=float(money(row.get("Bruto",0))), step=0.01, format="%.2f", key=f"bruto_{year}_{month}")
    ss_pct=c2.number_input("SS %", value=float((money(row.get("SS",0))/bruto*100) if bruto and money(row.get("SS",0)) else 4.85), step=0.01, format="%.2f")
    desempleo_pct=c3.number_input("Desempleo %", value=float((money(row.get("Desempleo",0))/bruto*100) if bruto and money(row.get("Desempleo",0)) else 1.65), step=0.01, format="%.2f")
    irpf_pct=c4.number_input("IRPF %", value=float((money(row.get("IRPF_pct",0)) or 0.13)*100), step=0.01, format="%.2f")
    ingresado=c5.number_input("Ingresado", value=float(money(row.get("Ingresado",0))), step=0.01, format="%.2f")
    ss=bruto*ss_pct/100; desemp=bruto*desempleo_pct/100; irpf=bruto*irpf_pct/100; neto=bruto-ss-desemp-irpf; dif=ingresado-neto if ingresado else 0
    estado="✅ Correcto" if ingresado and abs(dif)<0.02 else (f"🟢 Sobra {euro(dif)}" if dif>0 else (f"🔴 Falta {euro(abs(dif))}" if ingresado else "Pendiente"))
    k1,k2,k3,k4=st.columns(4); k1.metric("SS", euro(ss)); k2.metric("Desempleo", euro(desemp)); k3.metric("IRPF", euro(irpf)); k4.metric("Neto calculado", euro(neto), estado if ingresado else None)
    if st.button("Guardar nómina", use_container_width=True, key="save_nomina"):
        new={"Año":year,"Mes":month,"Fecha":date(year,month,1),"Bruto":bruto,"SS":ss,"Desempleo":desemp,"IRPF_pct":irpf_pct/100,"IRPF":irpf,"Neto_calculado":neto,"Ingresado":ingresado,"Diferencia":dif,"Estado":estado}
        df2=df[~((pd.to_numeric(df.get("Año"), errors="coerce")==year)&(pd.to_numeric(df.get("Mes"), errors="coerce")==month))].copy() if not df.empty else pd.DataFrame()
        save_table("App_Nominas", pd.concat([df2,pd.DataFrame([new])], ignore_index=True)); st.success("Nómina guardada."); st.rerun()
    ydf=df[pd.to_numeric(df.get("Año"), errors="coerce")==year].copy() if not df.empty else pd.DataFrame()
    if not ydf.empty:
        show=ydf.sort_values("Mes", ascending=False).copy(); show["Mes_txt"]=show.apply(lambda r:f"{MONTHS_ES[int(r['Mes'])-1]}{str(int(r['Año']))[-2:]}", axis=1)
        st.dataframe(dataframe_money(show[["Mes_txt","Bruto","SS","Desempleo","IRPF","Neto_calculado","Ingresado","Diferencia","Estado"]], ["Bruto","SS","Desempleo","IRPF","Neto_calculado","Ingresado","Diferencia"]), use_container_width=True, hide_index=True)
        a,b,c,d=st.columns(4); a.metric("Bruto acumulado", euro(ydf["Bruto"].apply(money).sum())); b.metric("Retención IRPF", euro(ydf["IRPF"].apply(money).sum())); c.metric("Gastos deducibles", euro(ydf["SS"].apply(money).sum()+ydf["Desempleo"].apply(money).sum())); d.metric("Neto acumulado", euro(ydf["Neto_calculado"].apply(money).sum()))
    render_vacaciones(year)

def load_vacaciones(): return load_table("App_Vacaciones", ["Año","Inicio","Fin","Dias","Notas"])

def render_month_calendar(year, month, festivos, vac_days):
    cal=calendar.Calendar(firstweekday=0)
    html='<div class="cal-grid">'+''.join([f'<div class="cal-head">{d}</div>' for d in ["L","M","X","J","V","S","D"]])
    for week in cal.monthdatescalendar(year, month):
        for d in week:
            if d.month!=month: cls="cal-day cal-empty"; txt=""
            elif d in vac_days: cls="cal-day cal-green"; txt=str(d.day)
            elif d in festivos: cls="cal-day cal-maroon"; txt=str(d.day)
            elif d.weekday()>=5: cls="cal-day cal-red"; txt=str(d.day)
            elif is_intensiva(d, festivos): cls="cal-day cal-grey"; txt=str(d.day)
            else: cls="cal-day cal-normal"; txt=str(d.day)
            html += f'<div class="{cls}">{txt}</div>'
    html+='</div>'; st.markdown(html, unsafe_allow_html=True)

def calendar_picker(label, year, key_prefix, min_date=None, default=None):
    if default is None: default = min_date or (date.today() if date.today().year==year else date(year,1,1))
    if min_date and default < min_date: default = min_date
    month_default=default.month
    m=st.selectbox(f"Mes {label}", list(range(1,13)), index=month_default-1, format_func=lambda x:f"{MONTHS_ES[x-1]} {year}", key=f"{key_prefix}_month")
    first=date(year,m,1); _,last_day=calendar.monthrange(year,m); last=date(year,m,last_day)
    start=max(first, min_date) if min_date else first
    days=[d for d in date_range(start,last)]
    if not days:
        st.info("No hay días disponibles en este mes desde la fecha inicio.")
        return min_date or first
    options=[None]+days
    def fmt(d): return "Seleccionar..." if d is None else f"{['Lun','Mar','Mié','Jue','Vie','Sáb','Dom'][d.weekday()]} {d.strftime('%d/%m/%Y')}"
    idx=options.index(default) if default in options else 0
    selected=st.selectbox(label, options, index=idx, format_func=fmt, key=f"{key_prefix}_day")
    return selected

def render_festivos_editor(year):
    with st.expander("🟤 Revisar / modificar festivos", expanded=False):
        fdf=load_festivos_df(year).copy(); fdf["Fecha"]=pd.to_datetime(fdf["Fecha"], errors="coerce").dt.date; fdf["Activo"]=fdf["Activo"].apply(boolify)
        edited=st.data_editor(fdf[["Fecha","Nombre","Activo","Origen"]].sort_values("Fecha"), num_rows="dynamic", use_container_width=True, hide_index=True, key=f"festivos_{year}", column_config={"Fecha":st.column_config.DateColumn("Fecha", format="DD/MM/YYYY"), "Activo":st.column_config.CheckboxColumn("Activo")})
        c1,c2=st.columns(2)
        if c1.button("Guardar festivos", use_container_width=True, key=f"save_fest_{year}"):
            edited["Año"]=year; save_festivos_year(year, edited[["Año","Fecha","Nombre","Activo","Origen"]]); st.success("Festivos guardados."); st.rerun()
        if c2.button("Restaurar base", use_container_width=True, key=f"reset_fest_{year}"):
            save_festivos_year(year, festivos_vitoria_base(year)); st.success("Restaurados."); st.rerun()

def render_vacaciones(year):
    st.divider(); st.subheader("🌴 Vacaciones y calendario laboral")
    render_festivos_editor(year); festivos=festivos_vitoria(year); df=load_vacaciones()
    if not df.empty: df["Año"]=pd.to_numeric(df["Año"], errors="coerce").fillna(year).astype(int)
    ydf=df[df["Año"]==year].copy() if not df.empty else pd.DataFrame(columns=["Año","Inicio","Fin","Dias","Notas"])
    used=sum(money(x) for x in ydf.get("Dias",[])); c1,c2,c3=st.columns(3); c1.metric("Días disponibles", VACACIONES_ANUALES); c2.metric("Días usados", f"{used:g}"); c3.metric("Días restantes", f"{VACACIONES_ANUALES-used:g}")
    with st.expander("➕ Añadir periodo de vacaciones", expanded=False):
        st.caption("Selector mensual con semanas de lunes a domingo. La fecha fin solo permite días desde la fecha inicio.")
        start=calendar_picker("Fecha inicio", year, "vac_start")
        if start:
            end=calendar_picker("Fecha fin", year, "vac_end", min_date=start, default=start)
        else: end=None
        calc=working_vacation_days(start,end,festivos) if start and end else 0
        dias=st.number_input("Días calculados/editables", value=float(calc), step=0.5, format="%.1f", key="vac_dias")
        notas=st.text_input("Notas", value="", key="vac_notas")
        if st.button("Guardar vacaciones", use_container_width=True, key="save_vac"):
            if not start or not end: st.error("Elige inicio y fin.")
            else:
                new={"Año":year,"Inicio":start.isoformat(),"Fin":end.isoformat(),"Dias":dias,"Notas":notas}
                save_table("App_Vacaciones", pd.concat([df,pd.DataFrame([new])], ignore_index=True) if not df.empty else pd.DataFrame([new])); st.success("Vacaciones guardadas."); st.rerun()
    if not ydf.empty:
        ydf_edit=ydf.copy(); ydf_edit["Borrar"]=False
        edited=st.data_editor(ydf_edit[["Borrar","Inicio","Fin","Dias","Notas"]], use_container_width=True, hide_index=True, key=f"vac_editor_{year}", column_config={"Borrar":st.column_config.CheckboxColumn("Borrar"), "Inicio":st.column_config.DateColumn("Inicio", format="DD/MM/YYYY"), "Fin":st.column_config.DateColumn("Fin", format="DD/MM/YYYY"), "Dias":st.column_config.NumberColumn("Días", step=.5)})
        if st.button("Guardar cambios vacaciones", use_container_width=True, key=f"save_vac_editor_{year}"):
            other=df[df["Año"]!=year].copy() if not df.empty else pd.DataFrame(columns=["Año","Inicio","Fin","Dias","Notas"])
            new=edited[~edited["Borrar"].fillna(False)].drop(columns=["Borrar"]); new["Año"]=year
            save_table("App_Vacaciones", pd.concat([other,new[["Año","Inicio","Fin","Dias","Notas"]]], ignore_index=True)); st.success("Vacaciones actualizadas."); st.rerun()
    vac_days=set()
    for _,r in ydf.iterrows():
        try: vac_days.update(date_range(pd.to_datetime(r["Inicio"]).date(), pd.to_datetime(r["Fin"]).date()))
        except Exception: pass
    cols=st.columns(3)
    for m in range(1,13):
        with cols[(m-1)%3]:
            st.markdown(f"**{MONTHS_ES[m-1]} {year}**"); render_month_calendar(year,m,festivos,vac_days)

# ---------- intereses ----------
def migrate_old_intereses_to_generic(df):
    if df.empty: return pd.DataFrame(columns=["Año","Mes","Fecha","Banco","Interes_bruto","Saldo","Retencion","Neto_esperado","Ingresado","Diferencia","Estado"])
    if "Banco" in df.columns: return df
    rows=[]
    for _,r in df.iterrows():
        for idx, banco in [(1,"BBVA"),(2,"Openbank")]:
            bruto=money(r.get(f"Cuenta{idx}_Interes",0)); saldo=money(r.get(f"Cuenta{idx}_Saldo",0))
            if bruto==0 and saldo==0: continue
            ret=bruto*RETENCION_INTERESES; neto=bruto-ret
            rows.append({"Año":r.get("Año"),"Mes":r.get("Mes"),"Fecha":r.get("Fecha"),"Banco":banco,"Interes_bruto":bruto,"Saldo":saldo,"Retencion":ret,"Neto_esperado":neto,"Ingresado":0,"Diferencia":0,"Estado":"Migrado"})
    return pd.DataFrame(rows)

def load_intereses_generic():
    headers=["Año","Mes","Fecha","Banco","Interes_bruto","Saldo","Retencion","Neto_esperado","Ingresado","Diferencia","Estado"]
    df=load_table("App_Intereses", headers)
    if not df.empty and "Banco" in df.columns: return df
    old_headers=["Año","Mes","Fecha","Cuenta1_Interes","Cuenta1_Saldo","Cuenta2_Interes","Cuenta2_Saldo","Retencion","Neto_esperado","Ingresado","Diferencia","Estado"]
    old=load_table("App_Intereses", old_headers)
    migrated=migrate_old_intereses_to_generic(old)
    if not migrated.empty: save_table("App_Intereses", migrated[headers])
    return migrated if not migrated.empty else pd.DataFrame(columns=headers)

def save_intereses(df):
    headers=["Año","Mes","Fecha","Banco","Interes_bruto","Saldo","Retencion","Neto_esperado","Ingresado","Diferencia","Estado"]
    out=df.copy()
    for h in headers:
        if h not in out: out[h]=None
    out["Año"]=pd.to_numeric(out["Año"], errors="coerce").fillna(date.today().year).astype(int)
    out["Mes"]=pd.to_numeric(out["Mes"], errors="coerce").fillna(1).astype(int)
    out["Fecha"]=out.apply(lambda r: date(int(r["Año"]), int(r["Mes"]), 1), axis=1)
    out["Banco"]=out["Banco"].fillna("Otros").astype(str).apply(safe_key)
    out["Interes_bruto"]=out["Interes_bruto"].apply(money); out["Saldo"]=out["Saldo"].apply(money); out["Ingresado"]=out["Ingresado"].apply(money)
    out["Retencion"]=out["Interes_bruto"]*RETENCION_INTERESES; out["Neto_esperado"]=out["Interes_bruto"]-out["Retencion"]; out["Diferencia"]=out["Ingresado"]-out["Neto_esperado"]
    out["Estado"]=out.apply(lambda r: "✅ Correcto" if r["Ingresado"] and abs(r["Diferencia"])<0.02 else (f"🟢 Sobra {euro(r['Diferencia'])}" if r["Diferencia"]>0 and r["Ingresado"] else (f"🔴 Falta {euro(abs(r['Diferencia']))}" if r["Ingresado"] else "Pendiente")), axis=1)
    save_table("App_Intereses", out[headers])

def render_intereses():
    st.header("🏦 Intereses cuentas")
    render_bank_config("intereses")
    df=load_intereses_generic(); years=list(range(2023, date.today().year+8)); year=st.selectbox("Año", years, index=years.index(date.today().year), key="int_year")
    keys=all_bank_keys(False)
    with st.expander("➕ Añadir interés mensual", expanded=False):
        month=st.selectbox("Mes", list(range(1,13)), index=date.today().month-1, format_func=lambda m:f"{MONTHS_ES[m-1]}{str(year)[-2:]}", key="int_add_month")
        banco=st.selectbox("Banco", keys or all_bank_keys(True), format_func=bank_name, key="int_add_banco")
        c1,c2,c3=st.columns(3)
        bruto=c1.number_input("Interés bruto", step=0.01, format="%.2f", key="int_bruto")
        saldo=c2.number_input("Saldo", step=0.01, format="%.2f", key="int_saldo")
        ingresado=c3.number_input("Ingresado", step=0.01, format="%.2f", key="int_ing")
        ret=bruto*RETENCION_INTERESES; neto=bruto-ret; dif=ingresado-neto if ingresado else 0
        a,b,c=st.columns(3); a.metric("Retención 19%", euro(ret)); b.metric("Neto esperado", euro(neto)); c.metric("Diferencia", euro(dif))
        if st.button("Guardar interés", use_container_width=True, key="save_int_add"):
            new={"Año":year,"Mes":month,"Fecha":date(year,month,1),"Banco":banco,"Interes_bruto":bruto,"Saldo":saldo,"Retencion":ret,"Neto_esperado":neto,"Ingresado":ingresado,"Diferencia":dif,"Estado":""}
            save_intereses(pd.concat([df,pd.DataFrame([new])], ignore_index=True) if not df.empty else pd.DataFrame([new])); st.success("Interés guardado."); st.rerun()
    ydf=df[pd.to_numeric(df.get("Año"), errors="coerce")==year].copy() if not df.empty else pd.DataFrame()
    if not ydf.empty:
        ydf["Mes_txt"]=ydf.apply(lambda r:f"{MONTHS_ES[int(r['Mes'])-1]}{str(int(r['Año']))[-2:]}", axis=1); ydf["Banco_nombre"]=ydf["Banco"].apply(bank_name)
        show=ydf.sort_values(["Mes","Banco"], ascending=[False, True])[["Mes_txt","Banco_nombre","Interes_bruto","Saldo","Retencion","Neto_esperado","Ingresado","Diferencia","Estado"]]
        st.dataframe(dataframe_money(show,["Interes_bruto","Saldo","Retencion","Neto_esperado","Ingresado","Diferencia"]), use_container_width=True, hide_index=True)
        st.subheader("Totales del año")
        a,b,c=st.columns(3); a.metric("Intereses brutos", euro(ydf["Interes_bruto"].apply(money).sum())); b.metric("Retenciones", euro(ydf["Retencion"].apply(money).sum())); c.metric("Neto esperado", euro(ydf["Neto_esperado"].apply(money).sum()))
    with st.expander("✏️ Editar o borrar intereses", expanded=False):
        edit=ydf.copy() if not ydf.empty else pd.DataFrame(columns=["Año","Mes","Banco","Interes_bruto","Saldo","Ingresado"])
        edit["Borrar"]=False
        edited=st.data_editor(edit[["Borrar","Año","Mes","Banco","Interes_bruto","Saldo","Ingresado"]], num_rows="dynamic", use_container_width=True, hide_index=True, key="editor_intereses", column_config={"Borrar":st.column_config.CheckboxColumn("Borrar"), "Mes":st.column_config.NumberColumn("Mes", min_value=1, max_value=12, step=1), "Banco":st.column_config.SelectboxColumn("Banco", options=all_bank_keys(True), format_func=bank_name), "Interes_bruto":st.column_config.NumberColumn("Interés bruto", step=0.01, format="%.2f"), "Saldo":st.column_config.NumberColumn("Saldo", step=0.01, format="%.2f"), "Ingresado":st.column_config.NumberColumn("Ingresado", step=0.01, format="%.2f")})
        if st.button("Guardar cambios intereses", use_container_width=True, key="save_int_editor"):
            other=df[pd.to_numeric(df.get("Año"), errors="coerce")!=year].copy() if not df.empty else pd.DataFrame()
            new=edited[~edited["Borrar"].fillna(False)].drop(columns=["Borrar"], errors="ignore")
            save_intereses(pd.concat([other,new], ignore_index=True)); st.success("Intereses actualizados."); st.rerun()

# ---------- IRPF ----------
def cuota_general_alava(base):
    # Escala orientativa editable en tabla; mantiene los tramos solicitados por capturas si se ajustan después.
    tramos=[(0,0,0.23),(17710,4073.30,0.28),(35420,9032.10,0.35),(53130,15230.60,0.40),(70840,22314.60,0.45),(106260,38253.60,0.46),(141680,54546.80,0.47),(204070,83870.10,0.49)]
    base=max(base,0)
    applicable=tramos[0]
    for t in tramos:
        if base>=t[0]: applicable=t
    return applicable[1]+(base-applicable[0])*applicable[2]

def reduccion_trabajo_alava(rend):
    if rend<=14800: return 8000
    if rend<=23000: return max(0, 8000-0.6098*(rend-14800))
    return 0

def cuota_ahorro_alava(base):
    base=max(base,0)
    if base<=2500: return base*.20
    if base<=10000: return 500+(base-2500)*.21
    if base<=15000: return 2075+(base-10000)*.22
    if base<=30000: return 3175+(base-15000)*.23
    return 6625+(base-30000)*.25

def load_irpf_manual(year):
    return load_table("App_IRPF_Manual", ["Año","Concepto","Importe","Editable"])

def render_irpf():
    st.header("📄 IRPF")
    years=list(range(2023, date.today().year+8)); year=st.selectbox("Año IRPF", years, index=years.index(date.today().year), key="irpf_year")
    nom=load_nominas(); inte=load_intereses_generic()
    ny=nom[pd.to_numeric(nom.get("Año"), errors="coerce")==year] if not nom.empty else pd.DataFrame()
    iy=inte[pd.to_numeric(inte.get("Año"), errors="coerce")==year] if not inte.empty else pd.DataFrame()
    rend_int=ny["Bruto"].apply(money).sum() if not ny.empty else 0
    gastos=(ny["SS"].apply(money).sum()+ny["Desempleo"].apply(money).sum()) if not ny.empty else 0
    rend_pre=max(rend_int-gastos,0); reduccion=reduccion_trabajo_alava(rend_pre); base_general=max(rend_pre-reduccion,0)
    cap_mob=iy["Interes_bruto"].apply(money).sum() if not iy.empty else 0
    escala_general=cuota_general_alava(base_general); minoracion=1583; cuota_general=max(escala_general-minoracion,0); cuota_ahorro=cuota_ahorro_alava(cap_mob)
    ret_trabajo=ny["IRPF"].apply(money).sum() if not ny.empty else 0; ret_ahorro=iy["Retencion"].apply(money).sum() if not iy.empty else 0
    total_cuota=cuota_general+cuota_ahorro; pagos=ret_trabajo+ret_ahorro; diferencial=total_cuota-pagos
    auto=pd.DataFrame([
        {"Concepto":"Rendimientos íntegros del trabajo", "Importe":rend_int, "Origen":"Nóminas brutas"},
        {"Concepto":"Gastos deducibles", "Importe":gastos, "Origen":"SS + desempleo"},
        {"Concepto":"Reducción/bonificación trabajo", "Importe":reduccion, "Origen":"Fórmula Álava"},
        {"Concepto":"Base liquidable general", "Importe":base_general, "Origen":"Automático"},
        {"Concepto":"Rendimiento capital mobiliario", "Importe":cap_mob, "Origen":"Intereses brutos"},
        {"Concepto":"Resultado escala general", "Importe":escala_general, "Origen":"Tramos Álava"},
        {"Concepto":"Minoración", "Importe":minoracion, "Origen":"Editable"},
        {"Concepto":"Cuota íntegra general", "Importe":cuota_general, "Origen":"Escala - minoración"},
        {"Concepto":"Cuota íntegra ahorro", "Importe":cuota_ahorro, "Origen":"Escala ahorro"},
        {"Concepto":"Retenciones trabajo", "Importe":ret_trabajo, "Origen":"Nóminas"},
        {"Concepto":"Retenciones capital mobiliario", "Importe":ret_ahorro, "Origen":"Intereses"},
        {"Concepto":"Cuota diferencial", "Importe":diferencial, "Origen":"Cuota - pagos"},
    ])
    manual=load_irpf_manual(year)
    if manual.empty or manual[pd.to_numeric(manual.get("Año"), errors="coerce")==year].empty:
        editable=auto.copy(); editable["Año"]=year; editable["Editable"]=True
    else:
        editable=manual[pd.to_numeric(manual.get("Año"), errors="coerce")==year].copy()
    st.caption("Campos automáticos, pero editables. Puedes guardar cambios manuales o restaurar los cálculos automáticos.")
    edited=st.data_editor(editable[["Concepto","Importe","Origen"]] if "Origen" in editable else editable[["Concepto","Importe"]], use_container_width=True, hide_index=True, num_rows="fixed", key=f"irpf_editor_{year}", column_config={"Importe":st.column_config.NumberColumn("Importe", step=0.01, format="%.2f")})
    c1,c2=st.columns(2)
    if c1.button("Guardar IRPF manual", use_container_width=True, key="save_irpf_manual"):
        cur=load_irpf_manual(year); other=cur[pd.to_numeric(cur.get("Año"), errors="coerce")!=year].copy() if not cur.empty else pd.DataFrame(columns=["Año","Concepto","Importe","Editable"])
        out=edited.copy(); out["Año"]=year; out["Editable"]=True
        save_table("App_IRPF_Manual", pd.concat([other,out[["Año","Concepto","Importe","Editable"]]], ignore_index=True)); st.success("IRPF manual guardado."); st.rerun()
    if c2.button("Restaurar cálculos automáticos", use_container_width=True, key="reset_irpf"):
        cur=load_irpf_manual(year); other=cur[pd.to_numeric(cur.get("Año"), errors="coerce")!=year].copy() if not cur.empty else pd.DataFrame(columns=["Año","Concepto","Importe","Editable"])
        save_table("App_IRPF_Manual", other); st.success("Cálculos restaurados."); st.rerun()
    res_color="normal" if diferencial<=0 else "inverse"
    st.metric("Resultado estimado", euro(diferencial), "A devolver" if diferencial<0 else "A pagar" if diferencial>0 else "Cero")

# ---------- header/footer ----------
def render_header():
    src=image_html(MFE_LOGO)
    logo=f'<img src="{src}" alt="MFE">' if src else ''
    st.markdown(f'<div class="header-brand">{logo}<div><h1 style="margin-bottom:0">💰 Ahorro Mikel</h1><div style="opacity:.75">Patrimonio · Nóminas · Vacaciones · Intereses · IRPF</div></div></div>', unsafe_allow_html=True)

def render_footer():
    src=image_html(MFE_LOGO)
    logo=f'<img src="{src}" alt="MFE">' if src else ''
    st.markdown(f'<div class="footer">{logo}<span><b>Ahorro Mikel</b> v{APP_VERSION}</span><span>Actualizada: {APP_UPDATED}</span><span>Usuario: mikelferech</span></div>', unsafe_allow_html=True)

# ---------- main ----------
login_gate()
col_title, col_out = st.columns([5,1])
with col_title: render_header()
with col_out:
    if st.button("Cerrar sesión", use_container_width=True): st.session_state.auth_ok=False; st.rerun()

if EXCEL_PATH.exists():
    with open(EXCEL_PATH,"rb") as f:
        st.download_button("⬇️ Descargar Excel actualizado", f, file_name="Ahorro_actualizado.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

tabs=st.tabs(["Dashboard","Ahorro","Nóminas","Intereses","IRPF"])
with tabs[0]:
    st.header("📊 Dashboard")
    df=load_app_ahorro(); kpi_dashboard(df)
    if not df.empty:
        chart=df.sort_values("Fecha").copy(); chart["Mes"]=chart["Fecha"].apply(mes_label)
        c1,c2=st.columns(2)
        with c1:
            fig=go.Figure(go.Scatter(x=chart["Mes"], y=chart["Total"], mode="lines+markers")); fig.update_layout(title="Patrimonio histórico", height=360); st.plotly_chart(fig, use_container_width=True)
        with c2:
            colors=["#16a34a" if v>=0 else "#dc2626" for v in chart["+/-"]]; fig2=go.Figure(go.Bar(x=chart["Mes"], y=chart["+/-"], marker_color=colors)); fig2.update_layout(title="+/- mensual", height=360); st.plotly_chart(fig2, use_container_width=True)
with tabs[1]: render_ahorro()
with tabs[2]: render_nominas()
with tabs[3]: render_intereses()
with tabs[4]: render_irpf()
render_footer()
